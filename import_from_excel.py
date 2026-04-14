"""
Import issues from the Gitea meeting Excel control table (CLI).

Usage:
    python import_from_excel.py --file "samples/gitea table sample.xlsx"

Idempotent: re-running will UPDATE existing issues (matched by display_number).

Note: For ongoing updates after initial import, use the web-based
Admin -> Excel Update feature instead, which provides diff preview
and conflict detection.
"""
import argparse
import io
import re
import sys
from pathlib import Path

# Ensure stdout can handle Unicode (Chinese characters, etc.)
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent))

from app import create_app
from app.db import get_db
from app.models import node as node_model
from app.models import user as user_model

# ---------- State parsing ----------

STATE_MAP = {
    "done": "done",
    "uat done": "uat_done",
    "uatdone": "uat_done",
    "uat": "uat",
    "developing": "developing",
    "dev": "developing",
    "tbd": "tbd",
    "unneeded": "unneeded",
}

# Excel header → node code mapping
HEADER_TO_CODE = {
    "A10":    "n_a10",
    "A12":    "n_a12",
    "A14":    "n_a14",
    "N2":     "n_n2",
    "A16":    "n_a16",
    "N2/A16": ["n_n2", "n_a16"],   # combined column → import to both nodes
    "N3":     "n_n3",
    "N4/N5":  "n_n4n5",
    "N6/N7":  "n_n6n7",
    "000":    "n_000",
    "MtM":    "n_mtm",
}

WK_PATTERN = re.compile(r"^wk(\d+)$", re.IGNORECASE)
DATE_PATTERN = re.compile(r"(\d{1,2})/(\d{1,2})")


def parse_cell(raw_value):
    """Parse a cell value into (state, check_in_date, short_note).

    Examples:
        "Done"                         -> ("done", None, None)
        "UAT done\\n2/20 Check in"     -> ("uat_done", "02-20", None)
        "Developing added spec for N4" -> ("developing", None, "added spec for N4")
        ""  / None                     -> (None, None, None)
    """
    if raw_value is None:
        return None, None, None

    text = str(raw_value).strip()
    if not text:
        return None, None, None

    # Split on newline — first part is state, rest may contain date/note
    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]

    state = None
    check_in_date = None
    short_note = None

    if not lines:
        return None, None, None

    # Try to match the first line (or beginning of first line) to a known state
    first_line = lines[0]
    first_lower = first_line.lower()

    # Check exact match first
    if first_lower in STATE_MAP:
        state = STATE_MAP[first_lower]
    else:
        # Try prefix match: "Developing added spec for N4 to N7"
        for key in sorted(STATE_MAP.keys(), key=len, reverse=True):
            if first_lower.startswith(key):
                state = STATE_MAP[key]
                remainder = first_line[len(key):].strip()
                if remainder:
                    short_note = remainder
                break

    if state is None:
        # Could not parse state — store entire text as short_note
        short_note = text
        return None, None, short_note

    # Check remaining lines for date
    for line in lines[1:]:
        date_match = DATE_PATTERN.search(line)
        if date_match:
            month = int(date_match.group(1))
            day = int(date_match.group(2))
            check_in_date = f"{month:02d}-{day:02d}"
        else:
            # Append to short_note
            if short_note:
                short_note += " " + line
            else:
                short_note = line

    return state, check_in_date, short_note


def expand_merged_cells(ws):
    """For each merged range, fill all cells with the top-left value."""
    for merge_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merge_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


def import_sheet(ws, node_lookup, legacy_user_id, is_closed_sheet=False):
    """Import one worksheet. Returns (imported_count, skipped_count, errors)."""
    expand_merged_cells(ws)

    # --- Parse header row ---
    header_row = []
    for cell in ws[1]:
        header_row.append(cell.value)

    # Build column index: col_number -> node_id
    node_columns = {}
    owner_col = None
    status_col = None
    jira_col = None
    icv_col = None
    uat_path_col = None
    topic_col = None

    for idx, header in enumerate(header_row):
        if header is None:
            continue
        h = str(header).strip()
        if h == "Status":
            status_col = idx
        elif h == "Owner":
            owner_col = idx
        elif h in ("JIRA", "JIRA No.", "JIRA No"):
            jira_col = idx
        elif h == "ICV":
            icv_col = idx
        elif h in ("UAT path", "UAT Path", "UATpath"):
            uat_path_col = idx
        elif h == "Topic":
            topic_col = idx
        elif h in HEADER_TO_CODE:
            codes = HEADER_TO_CODE[h]
            if isinstance(codes, str):
                codes = [codes]
            for code in codes:
                if code in node_lookup:
                    node_columns.setdefault(idx, []).append(node_lookup[code])

    if not node_columns:
        return 0, 0, ["No node columns found in header"]

    db = get_db()
    imported = 0
    skipped = 0
    errors = []
    current_week_year = None
    current_week_number = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        first_val = row[0].value
        if first_val is None:
            continue

        first_str = str(first_val).strip()
        if not first_str:
            continue

        # Check if it's a week separator
        wk_match = WK_PATTERN.match(first_str)
        if wk_match:
            raw_week = int(wk_match.group(1))
            # wk321 → week 321 mod 100? No — these are literal week numbers.
            # But ISO weeks only go to 52/53. If > 100 it might be year-encoded.
            # For sample data: wk321 likely means week 21 of some year, or just ordinal.
            # We'll store as-is since we don't know the year context.
            # Convention: if <= 53, assume current context year; if > 100, parse as YYww
            if raw_week > 100:
                # e.g., 321 -> year_offset=3, week=21 -> 2023 wk21
                # or 325 -> year_offset=3, week=25 -> 2023 wk25
                year_part = raw_week // 100
                week_part = raw_week % 100
                # Assume 2020 + year_part as base
                current_week_year = 2020 + year_part
                current_week_number = week_part
            else:
                current_week_year = current_week_year or 2025
                current_week_number = raw_week
            continue

        # Try to parse as issue number
        try:
            display_number = str(int(float(first_str)))
        except (ValueError, TypeError):
            continue

        if current_week_year is None:
            current_week_year = 2025
            current_week_number = 1

        # Read cell values
        def cell_val(idx):
            if idx is not None and idx < len(row):
                return row[idx].value
            return None

        status_val = str(cell_val(status_col) or "ongoing").strip().lower()
        issue_status = "ongoing"
        if is_closed_sheet or status_val == "closed":
            issue_status = "closed"
        elif status_val == "on hold" or status_val == "on_hold":
            issue_status = "on_hold"

        owner_name = str(cell_val(owner_col) or "").strip() or None
        jira = str(cell_val(jira_col) or "").strip() or None
        icv = str(cell_val(icv_col) or "").strip() or None
        uat_path = str(cell_val(uat_path_col) or "").strip() or None
        topic = str(cell_val(topic_col) or "").strip() or f"Issue #{display_number}"

        # Upsert issue
        existing = db.execute(
            "SELECT id FROM issues WHERE display_number = ?",
            (display_number,),
        ).fetchone()

        now_str = __import__("datetime").datetime.now(
            __import__("datetime").timezone.utc
        ).isoformat()

        if existing:
            issue_id = existing["id"]
            db.execute(
                """UPDATE issues SET topic=?, requestor_name=?, week_year=?,
                   week_number=?, jira_ticket=?, icv=?, uat_path=?, status=?,
                   updated_at=?
                   WHERE id=?""",
                (topic, owner_name, current_week_year, current_week_number,
                 jira, icv, uat_path, issue_status, now_str, issue_id),
            )
        else:
            cur = db.execute(
                """INSERT INTO issues
                   (display_number, topic, requestor_name, owner_user_id,
                    week_year, week_number, jira_ticket, icv, uat_path,
                    status, created_at, updated_at, latest_update_at)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (display_number, topic, owner_name, None,
                 current_week_year, current_week_number, jira, icv, uat_path,
                 issue_status, now_str, now_str, now_str),
            )
            issue_id = cur.lastrowid

        # Process node states
        for col_idx, node_ids in node_columns.items():
            raw = cell_val(col_idx)
            state, check_in_date, short_note = parse_cell(raw)

            for node_id in node_ids:
                existing_state = db.execute(
                    "SELECT id FROM issue_node_states WHERE issue_id = ? AND node_id = ?",
                    (issue_id, node_id),
                ).fetchone()

                if existing_state:
                    db.execute(
                        """UPDATE issue_node_states
                           SET state=?, check_in_date=?, short_note=?,
                               updated_at=?, updated_by_user_id=?,
                               updated_by_name_snapshot=?
                           WHERE issue_id=? AND node_id=?""",
                        (state, check_in_date, short_note, now_str,
                         legacy_user_id, "Legacy", issue_id, node_id),
                    )
                else:
                    db.execute(
                        """INSERT INTO issue_node_states
                           (issue_id, node_id, state, check_in_date, short_note,
                            updated_at, updated_by_user_id, updated_by_name_snapshot)
                           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                        (issue_id, node_id, state, check_in_date, short_note,
                         now_str, legacy_user_id, "Legacy"),
                    )

        # Update cache
        cache = db.execute(
            """SELECT
                 MAX(updated_at) as latest,
                 MIN(CASE WHEN state IN ('done', 'unneeded') THEN 1
                          WHEN state IS NULL THEN 0
                          ELSE 0 END) as all_done
               FROM issue_node_states WHERE issue_id = ?""",
            (issue_id,),
        ).fetchone()
        db.execute(
            "UPDATE issues SET latest_update_at=?, all_nodes_done=? WHERE id=?",
            (cache["latest"], cache["all_done"] or 0, issue_id),
        )

        imported += 1

    db.commit()
    return imported, skipped, errors


def main():
    parser = argparse.ArgumentParser(description="Import Excel into Gitea Tracker")
    parser.add_argument("--file", required=True, help="Path to .xlsx file")
    args = parser.parse_args()

    xlsx_path = Path(args.file)
    if not xlsx_path.exists():
        print(f"ERROR: File not found: {xlsx_path}")
        sys.exit(1)

    app = create_app()
    with app.app_context():
        # Build node lookup: code -> node_id
        nodes = node_model.get_all_active()
        node_lookup = {n["code"]: n["id"] for n in nodes}
        if not node_lookup:
            print("ERROR: No nodes in DB. Run seed.py first.")
            sys.exit(1)

        # Get legacy user
        legacy = user_model.get_by_username("legacy")
        if not legacy:
            print("ERROR: Legacy user not found. Run seed.py first.")
            sys.exit(1)
        legacy_user_id = legacy["id"]

        wb = openpyxl.load_workbook(xlsx_path)
        total_imported = 0
        total_errors = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            is_closed = "closed" in sheet_name.lower()
            print(f"\nProcessing sheet: '{sheet_name}' (closed={is_closed})")

            imported, skipped, errors = import_sheet(
                ws, node_lookup, legacy_user_id, is_closed_sheet=is_closed
            )
            total_imported += imported
            total_errors.extend(errors)
            print(f"  Imported: {imported}, Errors: {len(errors)}")
            for err in errors:
                print(f"    - {err}")

        print(f"\n{'='*40}")
        print(f"Total imported: {total_imported}")
        if total_errors:
            print(f"Total errors: {len(total_errors)}")
        else:
            print("No errors.")


if __name__ == "__main__":
    main()
