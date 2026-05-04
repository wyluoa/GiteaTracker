"""
Excel export — round-trip the .xlsx through openpyxl and verify the locked
design choices. These tests are the regression net for state coloring,
red-line treatment, hyperlinks, and the two-sheet structure.
"""
import io

import pytest
from openpyxl import load_workbook

from app import excel_export


# ─── State color palette ──────────────────────────────────────────────

def test_state_font_uat_below_redline_is_orange():
    """UAT below the red line: standard orange (#D35400), bold."""
    f = excel_export.state_font("uat", above_red_line=False)
    assert f.color.rgb == "FFD35400"
    assert f.bold is True


def test_state_font_uat_above_redline_is_red():
    """LOCKED: above-red-line UAT must be red bold (overrides orange)."""
    f = excel_export.state_font("uat", above_red_line=True)
    assert f.color.rgb == excel_export.RED_LINE_COLOR
    assert f.bold is True


def test_state_font_tbd_above_redline_is_red():
    """LOCKED: TBD above red line is red too (UAT/TBD share treatment)."""
    f = excel_export.state_font("tbd", above_red_line=True)
    assert f.color.rgb == excel_export.RED_LINE_COLOR


def test_state_font_done_above_redline_unaffected():
    """Only UAT/TBD turn red above the red line. Done/Dev/etc. stay normal."""
    assert excel_export.state_font("done", above_red_line=True).color.rgb == "FF1E8449"
    assert excel_export.state_font("developing", above_red_line=True).color.rgb == "FF566573"


def test_state_font_uat_done_is_blue_not_red_above_redline():
    """uat_done is NOT in the {uat, tbd} red-line set — keeps its blue."""
    f = excel_export.state_font("uat_done", above_red_line=True)
    assert f.color.rgb == "FF2874A6"


def test_state_font_blank_returns_none():
    """Empty state → no Font (caller decides default)."""
    assert excel_export.state_font(None) is None
    assert excel_export.state_font("") is None


def test_above_red_line_boundary():
    """Boundary: red line wk426 → wk426 IS above (≤), wk427 is NOT."""
    assert excel_export.above_red_line(2024, 26, 2024, 26) is True
    assert excel_export.above_red_line(2024, 25, 2024, 26) is True
    assert excel_export.above_red_line(2024, 27, 2024, 26) is False
    # No red line set → never above
    assert excel_export.above_red_line(2024, 26, None, None) is False


# ─── Date formatting ──────────────────────────────────────────────────

def test_format_check_in_already_yyyy_mm_dd():
    """UI inputs YYYY-MM-DD; pass through untouched."""
    assert excel_export.format_check_in_date("2026-08-15", 2026, 26) == "2026-08-15"


def test_format_check_in_mm_dd_after_week_same_year():
    """8/15 after wk626 (2026-06-22) → same year."""
    assert excel_export.format_check_in_date("08-15", 2026, 26) == "2026-08-15"


def test_format_check_in_mm_dd_before_week_next_year():
    """2/20 before wk626 (2026-06-22) → next year."""
    assert excel_export.format_check_in_date("02-20", 2026, 26) == "2027-02-20"


def test_format_check_in_empty_returns_blank():
    assert excel_export.format_check_in_date(None, 2026, 26) == ""
    assert excel_export.format_check_in_date("", 2026, 26) == ""


def test_format_check_in_unknown_format_passes_through():
    """Don't lie about formats we don't recognize — pass through."""
    assert excel_export.format_check_in_date("Q3", 2026, 26) == "Q3"


# ─── Workbook structure ───────────────────────────────────────────────

def _build_test_workbook(*, ongoing=None, on_hold=None, closed=None,
                         nodes=None, all_states=None,
                         red_line=(2024, 26),
                         filtered=False,
                         gitea_url_for=None):
    return excel_export.build_workbook(
        ongoing_issues=ongoing or [],
        on_hold_issues=on_hold or [],
        closed_issues=closed or [],
        nodes=nodes or [],
        all_states=all_states or {},
        red_line_year=red_line[0] if red_line else None,
        red_line_week=red_line[1] if red_line else None,
        exporter_display_name="Test User",
        exporter_username="testuser",
        filtered=filtered,
        gitea_url_for=gitea_url_for,
    )


def _issue(**kwargs):
    base = {
        "id": 1, "display_number": "T001", "topic": "test",
        "requestor_name": "alice", "status": "ongoing",
        "week_year": 2026, "week_number": 26,
        "jira_ticket": None, "icv": None, "uat_path": None,
        "pending_close": 0, "group_label": None,
    }
    base.update(kwargs)
    return base


def test_workbook_has_two_sheets():
    """LOCKED A1: two sheets — Ongoing + Closed."""
    buf = _build_test_workbook()
    wb = load_workbook(buf)
    assert wb.sheetnames == ["Ongoing", "Closed"]


def test_theme_contains_calibri_and_jhenghei():
    """LOCKED font strategy: theme1.xml has Latin=Calibri + EA=Microsoft JhengHei."""
    import zipfile
    buf = _build_test_workbook()
    with zipfile.ZipFile(buf) as z:
        theme = z.read("xl/theme/theme1.xml").decode("utf-8")
    assert '<a:latin typeface="Calibri"' in theme
    assert "Microsoft JhengHei" in theme


def test_metadata_row_includes_exporter_and_redline():
    """Row 1 metadata band: exporter name, red line label, range, count."""
    buf = _build_test_workbook(
        ongoing=[_issue(display_number="A", week_year=2026, week_number=26)],
        red_line=(2024, 26),
    )
    wb = load_workbook(buf)
    meta = wb["Ongoing"].cell(1, 1).value
    assert "Test User" in meta
    assert "wk426" in meta
    assert "全部" in meta
    assert "題目數 1" in meta


def test_metadata_says_filtered_when_filtered():
    """When filtered=True, metadata says '目前篩選結果'."""
    buf = _build_test_workbook(filtered=True)
    wb = load_workbook(buf)
    assert "目前篩選結果" in wb["Ongoing"].cell(1, 1).value


def test_no_redline_metadata_dash():
    """Red line not configured → metadata shows '—'."""
    buf = _build_test_workbook(red_line=None)
    wb = load_workbook(buf)
    assert "紅線 —" in wb["Ongoing"].cell(1, 1).value


def test_status_friendly_label_with_pending_close_suffix():
    """LOCKED Q4: friendly Status; F2: pending_close adds '(待確認關單)'."""
    nodes = [{"id": 1, "display_name": "A10"}]
    states = {1: {1: {"state": "uat", "check_in_date": None, "short_note": None}}}
    buf = _build_test_workbook(
        ongoing=[
            _issue(id=1, display_number="N1", status="ongoing", pending_close=1),
            _issue(id=2, display_number="N2", status="on_hold"),
            _issue(id=3, display_number="N3", status="ongoing"),
        ],
        nodes=nodes,
        all_states={1: states[1], 2: {}, 3: {}},
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    # Find rows by display_number in col 1; statuses are in col 2.
    statuses = {ws.cell(r, 1).value: ws.cell(r, 2).value
                for r in range(4, ws.max_row + 1)}
    assert statuses["N1"] == "Ongoing (待確認關單)"
    assert statuses["N2"] == "On Hold"
    assert statuses["N3"] == "Ongoing"


def test_state_cell_shows_label_date_short_note_three_lines():
    """LOCKED B1=b: state cell is Label\\n日期\\nshort_note."""
    nodes = [{"id": 1, "display_name": "A10"}]
    states = {1: {1: {"state": "uat", "check_in_date": "2026-08-15",
                       "short_note": "等廠商回覆"}}}
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="X")],
        nodes=nodes, all_states=states,
        red_line=None,
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    # Header row 3, body starts row 4. With week separator, real first row is 5.
    # Find the cell with our state.
    state_cell = None
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "X":
            state_cell = ws.cell(r, 4)  # cols: # / Status / Owner / A10
            break
    assert state_cell is not None
    assert state_cell.value == "UAT\n2026-08-15\n等廠商回覆"


def test_state_cell_omits_blank_lines():
    """No check-in date and no short note → just the label."""
    nodes = [{"id": 1, "display_name": "A10"}]
    states = {1: {1: {"state": "done", "check_in_date": None, "short_note": None}}}
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="X")],
        nodes=nodes, all_states=states, red_line=None,
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "X":
            assert ws.cell(r, 4).value == "Done"
            return
    pytest.fail("X row not found")


def test_hash_column_hyperlink_when_resolver_returns_url():
    """LOCKED D1=a: # column is a hyperlink to Gitea when URL resolves."""
    def gitea_url(dn): return f"http://gitea/issues/{dn}"
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="123")],
        gitea_url_for=gitea_url,
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "123":
            assert ws.cell(r, 1).hyperlink is not None
            assert ws.cell(r, 1).hyperlink.target == "http://gitea/issues/123"
            return
    pytest.fail("issue 123 row not found")


def test_hash_column_no_hyperlink_when_resolver_blank():
    """If no URL mapping, # is plain text — don't fake a hyperlink."""
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="X")],
        gitea_url_for=lambda dn: "",
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "X":
            assert ws.cell(r, 1).hyperlink is None
            return
    pytest.fail("X row not found")


def test_above_redline_uat_cell_is_red_in_workbook():
    """End-to-end: an issue at-or-above red line with a UAT cell renders red."""
    nodes = [{"id": 1, "display_name": "A10"}]
    states = {1: {1: {"state": "uat", "check_in_date": None, "short_note": None}}}
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="OLD",
                        week_year=2024, week_number=10)],  # before red line
        nodes=nodes, all_states=states,
        red_line=(2024, 26),
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "OLD":
            assert ws.cell(r, 4).font.color.rgb == excel_export.RED_LINE_COLOR
            return
    pytest.fail("OLD row not found")


def test_freeze_panes_header_plus_first_three_cols():
    """LOCKED E1: pane freezes after row 3 (header band) and after col C."""
    buf = _build_test_workbook()
    wb = load_workbook(buf)
    assert wb["Ongoing"].freeze_panes == "D4"
    assert wb["Closed"].freeze_panes == "D4"


def _find_row_starting_with(ws, prefix):
    for r in range(4, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).startswith(prefix):
            return r
    return None


def test_wk_separator_row_fills_all_columns_gray():
    """Wk separator row gets a light-gray fill spanning every column,
    not just col A — so the visual band runs the full width of the table."""
    nodes = [{"id": 1, "display_name": "A10"}, {"id": 2, "display_name": "A12"}]
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="X", week_year=2026, week_number=26)],
        nodes=nodes, all_states={1: {}},
        red_line=None,
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    wk_row = _find_row_starting_with(ws, "wk")
    assert wk_row is not None, "wk separator row not found"
    total_cols = ws.max_column
    assert total_cols >= 5
    for col in range(1, total_cols + 1):
        rgb = ws.cell(wk_row, col).fill.start_color.rgb
        assert rgb == excel_export.WK_ROW_FILL_COLOR, (
            f"col {col} fill {rgb} != gray {excel_export.WK_ROW_FILL_COLOR}"
        )


def test_red_line_row_inserted_at_boundary():
    """When weeks cross the red line, an explicit divider row is inserted
    between the last 'above' wk-group and the first 'below' wk-group."""
    nodes = []
    buf = _build_test_workbook(
        ongoing=[
            _issue(id=1, display_number="OLD", week_year=2024, week_number=20),
            _issue(id=2, display_number="MID", week_year=2024, week_number=26),
            _issue(id=3, display_number="NEW", week_year=2024, week_number=30),
        ],
        nodes=nodes, all_states={1: {}, 2: {}, 3: {}},
        red_line=(2024, 26),
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    rl_row = _find_row_starting_with(ws, "─── redline")
    # (Label phrasing is intentionally neutral — no value judgment about
    # the items above the line.)
    assert rl_row is not None, "red-line marker row not found"
    # The marker should sit between MID (above) and NEW (below).
    mid_row = next(r for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value == "MID")
    new_row = next(r for r in range(4, ws.max_row + 1) if ws.cell(r, 1).value == "NEW")
    assert mid_row < rl_row < new_row
    # Fill is the soft-red color across all columns.
    for col in range(1, ws.max_column + 1):
        rgb = ws.cell(rl_row, col).fill.start_color.rgb
        assert rgb == excel_export.RED_LINE_ROW_FILL_COLOR


def test_no_red_line_row_when_all_above():
    """If every issue is above the red line there's no boundary to mark."""
    buf = _build_test_workbook(
        ongoing=[
            _issue(id=1, display_number="A", week_year=2024, week_number=10),
            _issue(id=2, display_number="B", week_year=2024, week_number=20),
        ],
        red_line=(2024, 26),
    )
    wb = load_workbook(buf)
    assert _find_row_starting_with(wb["Ongoing"], "─── 紅線") is None


def test_no_red_line_row_when_red_line_unset():
    """No red line configured → no marker row even with mixed weeks."""
    buf = _build_test_workbook(
        ongoing=[
            _issue(id=1, display_number="A", week_year=2024, week_number=10),
            _issue(id=2, display_number="B", week_year=2024, week_number=30),
        ],
        red_line=None,
    )
    wb = load_workbook(buf)
    assert _find_row_starting_with(wb["Ongoing"], "─── 紅線") is None


def test_owner_falls_back_to_requestor_name():
    """LOCKED 'matches tracker UI': Owner column = requestor_name verbatim."""
    nodes = []
    buf = _build_test_workbook(
        ongoing=[_issue(id=1, display_number="X", requestor_name="Bob")],
        nodes=nodes,
    )
    wb = load_workbook(buf)
    ws = wb["Ongoing"]
    for r in range(4, ws.max_row + 1):
        if ws.cell(r, 1).value == "X":
            assert ws.cell(r, 3).value == "Bob"
            return
    pytest.fail("X row not found")
