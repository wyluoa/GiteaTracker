"""
Excel export — builds the .xlsx workbook served by /export.

Design (locked 2026-04-29):
- Two sheets: Ongoing (incl. on_hold mixed in) + Closed
- Fonts via theme: Latin = Calibri, East Asian = Microsoft JhengHei
- State distinguished by FONT COLOR (no cell fill), reusing UI palette one
  shade deeper for white-background readability
- UAT/TBD above the red line are red+bold (covers the regular state color)
- Top metadata row + small legend so a recipient who never used the app can
  decode it without context
- # column hyperlinks into Gitea
- Frozen pane: row 1 of header band + first 3 columns
- Filename: gitea_tracker_YYYY-MM-DD_<username>.xlsx
"""
from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink


# ─── Theme: Calibri (Latin) + Microsoft JhengHei (East Asian) ──────────

# Inserted into xl/theme/theme1.xml. Cells that use scheme="minor" pick latin
# or ea per-character automatically — that's what gets us "English in Calibri,
# Chinese in 微軟正黑體" without per-cell font juggling.
_TAIWAN_THEME_XML = b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<a:theme xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" name="Gitea Tracker">
  <a:themeElements>
    <a:clrScheme name="Office">
      <a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>
      <a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>
      <a:dk2><a:srgbClr val="44546A"/></a:dk2>
      <a:lt2><a:srgbClr val="E7E6E6"/></a:lt2>
      <a:accent1><a:srgbClr val="4472C4"/></a:accent1>
      <a:accent2><a:srgbClr val="ED7D31"/></a:accent2>
      <a:accent3><a:srgbClr val="A5A5A5"/></a:accent3>
      <a:accent4><a:srgbClr val="FFC000"/></a:accent4>
      <a:accent5><a:srgbClr val="5B9BD5"/></a:accent5>
      <a:accent6><a:srgbClr val="70AD47"/></a:accent6>
      <a:hlink><a:srgbClr val="0563C1"/></a:hlink>
      <a:folHlink><a:srgbClr val="954F72"/></a:folHlink>
    </a:clrScheme>
    <a:fontScheme name="GiteaTracker">
      <a:majorFont>
        <a:latin typeface="Calibri Light"/>
        <a:ea typeface="Microsoft JhengHei"/>
        <a:cs typeface=""/>
      </a:majorFont>
      <a:minorFont>
        <a:latin typeface="Calibri"/>
        <a:ea typeface="Microsoft JhengHei"/>
        <a:cs typeface=""/>
      </a:minorFont>
    </a:fontScheme>
    <a:fmtScheme name="Office">
      <a:fillStyleLst>
        <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
        <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
        <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
      </a:fillStyleLst>
      <a:lnStyleLst>
        <a:ln w="6350" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>
        <a:ln w="12700" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>
        <a:ln w="19050" cap="flat" cmpd="sng" algn="ctr"><a:solidFill><a:schemeClr val="phClr"/></a:solidFill><a:prstDash val="solid"/></a:ln>
      </a:lnStyleLst>
      <a:effectStyleLst>
        <a:effectStyle><a:effectLst/></a:effectStyle>
        <a:effectStyle><a:effectLst/></a:effectStyle>
        <a:effectStyle><a:effectLst/></a:effectStyle>
      </a:effectStyleLst>
      <a:bgFillStyleLst>
        <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
        <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
        <a:solidFill><a:schemeClr val="phClr"/></a:solidFill>
      </a:bgFillStyleLst>
    </a:fmtScheme>
  </a:themeElements>
  <a:objectDefaults/>
  <a:extraClrSchemeLst/>
</a:theme>
"""


def apply_taiwan_theme(wb: Workbook) -> None:
    """Replace the workbook's default theme with one whose minor font scheme
    has Latin=Calibri + EastAsian=Microsoft JhengHei. Cells that use
    Font(scheme='minor') will then pick latin/ea per-character."""
    wb.loaded_theme = _TAIWAN_THEME_XML


# ─── State color palette + label map ──────────────────────────────────

# Same hues as the UI but deepened one shade for legibility on white.
# Bold for "needs attention" states; muted/italic for terminal/no-op states.
STATE_STYLES = {
    "done":       {"color": "FF1E8449", "bold": True,  "italic": False, "label": "Done"},
    "uat_done":   {"color": "FF2874A6", "bold": True,  "italic": False, "label": "UAT Done"},
    "uat":        {"color": "FFD35400", "bold": True,  "italic": False, "label": "UAT"},
    "tbd":        {"color": "FF7D3C98", "bold": True,  "italic": False, "label": "TBD"},
    "developing": {"color": "FF566573", "bold": False, "italic": False, "label": "Dev"},
    "unneeded":   {"color": "FF95A5A6", "bold": False, "italic": True,  "label": "Unneeded"},
}

# Color used for UAT/TBD cells in issues at-or-above the red line. Overrides
# the state color (still bold) so red-line stragglers leap out.
RED_LINE_COLOR = "FFC0392B"

# Friendly status labels for the issue-level Status column.
STATUS_LABELS = {
    "ongoing": "Ongoing",
    "on_hold": "On Hold",
    "closed":  "Closed",
}

# Header band style.
HEADER_COLOR = "FF1F3864"   # dark navy

# Metadata row style.
META_COLOR = "FF6C757D"     # muted gray

# Hyperlink style for the # column.
HYPERLINK_COLOR = "FF0563C1"

# Wk separator row — light gray fill across all columns.
WK_ROW_FILL_COLOR = "FFE8E8E8"

# Red-line marker row — soft red fill, bold dark-red text, spans all columns.
RED_LINE_ROW_FILL_COLOR = "FFFCE5E5"
RED_LINE_ROW_TEXT_COLOR = "FFC0392B"


# ─── Style builders ────────────────────────────────────────────────────

def header_font() -> Font:
    return Font(name="Calibri", scheme="minor", size=11, bold=True, color=HEADER_COLOR)


def header_border() -> Border:
    return Border(bottom=Side(style="thin", color="FFBFBFBF"))


def body_font_default() -> Font:
    return Font(scheme="minor", size=10)


def metadata_font(bold: bool = False) -> Font:
    return Font(scheme="minor", size=10, italic=True, color=META_COLOR, bold=bold)


def week_separator_font() -> Font:
    return Font(scheme="minor", size=11, bold=True, color="FF404040")


def week_separator_border() -> Border:
    return Border(top=Side(style="thin", color="FFBFBFBF"))


def week_separator_fill() -> PatternFill:
    return PatternFill(fill_type="solid", start_color=WK_ROW_FILL_COLOR,
                       end_color=WK_ROW_FILL_COLOR)


def red_line_row_font() -> Font:
    return Font(scheme="minor", size=11, bold=True, color=RED_LINE_ROW_TEXT_COLOR)


def red_line_row_fill() -> PatternFill:
    return PatternFill(fill_type="solid", start_color=RED_LINE_ROW_FILL_COLOR,
                       end_color=RED_LINE_ROW_FILL_COLOR)


def hyperlink_font() -> Font:
    return Font(scheme="minor", size=10, color=HYPERLINK_COLOR, underline="single")


def state_font(state: str | None, *, above_red_line: bool = False) -> Font | None:
    """Return Font for a state cell. None when the cell is blank (caller
    can decide to skip styling). Above-red-line UAT/TBD get the red color
    instead of their state hue (still bold)."""
    if not state:
        return None
    sty = STATE_STYLES.get(state)
    if not sty:
        return body_font_default()
    color = sty["color"]
    bold = sty["bold"]
    italic = sty["italic"]
    if above_red_line and state in ("uat", "tbd"):
        color = RED_LINE_COLOR
        bold = True
    return Font(scheme="minor", size=10, color=color, bold=bold, italic=italic)


# ─── Misc helpers ──────────────────────────────────────────────────────

def above_red_line(week_year: int, week_number: int,
                   red_line_year: int | None, red_line_week: int | None) -> bool:
    """An issue is at-or-above the red line when its week is older than (or
    equal to) the configured red-line week. Returns False if no red line."""
    if not red_line_year or not red_line_week:
        return False
    return (week_year, week_number) <= (red_line_year, red_line_week)


def format_check_in_date(raw: str | None,
                         issue_week_year: int | None,
                         issue_week_number: int | None) -> str:
    """Normalize a stored check_in_date to YYYY-MM-DD. Handles three formats:

      - YYYY-MM-DD  → returned as-is (UI input path)
      - MM-DD       → year derived via smart inference (Q1=b):
                       MM-DD on/after issue's ISO Monday → same year as issue
                       MM-DD before  issue's ISO Monday → next year
      - anything else / None → empty string

    Year inference is conservative: it favors "this is a future plan within
    the same year" over "last year's late milestone". For very old issues
    where this guess is wrong, the value can be edited via UI.
    """
    if not raw:
        return ""
    s = str(raw).strip()
    if not s:
        return ""
    # Already YYYY-MM-DD
    if len(s) == 10 and s[4] == "-" and s[7] == "-":
        return s
    # MM-DD
    if len(s) == 5 and s[2] == "-":
        try:
            mm = int(s[:2])
            dd = int(s[3:])
        except ValueError:
            return s
        if not (1 <= mm <= 12 and 1 <= dd <= 31):
            return s
        if issue_week_year and issue_week_number:
            year = _infer_year(mm, dd, issue_week_year, issue_week_number)
        else:
            year = date.today().year
        return f"{year:04d}-{mm:02d}-{dd:02d}"
    # Unknown format — pass through so we don't lie about it.
    return s


def _infer_year(mm: int, dd: int, week_year: int, week_number: int) -> int:
    """Q1=b smart year inference. If the MM-DD falls on or after the issue's
    ISO Monday → same year; before → next year (rolls into the next calendar
    year, matching "future plan" intuition)."""
    try:
        candidate = date(week_year, mm, dd)
    except ValueError:
        return week_year
    iso_monday = _iso_week_monday(week_year, week_number)
    return week_year if candidate >= iso_monday else week_year + 1


def _iso_week_monday(year: int, week: int) -> date:
    """Monday of the given ISO week. Pure-stdlib, avoids isocalendar weirdness
    by using fromisocalendar (Python 3.8+)."""
    try:
        return date.fromisocalendar(year, week, 1)
    except ValueError:
        return date(year, 1, 1)


# ─── Column layout ─────────────────────────────────────────────────────

# (label, width). Node columns are inserted between Owner and JIRA at runtime.
FIXED_LEAD = [("#", 8), ("Status", 12), ("Owner", 14)]
FIXED_TAIL = [("JIRA", 14), ("UAT Path", 30), ("Topic", 50), ("ICV", 14), ("Group Label", 18)]
NODE_WIDTH = 11


def _status_label(status: str, pending_close: int | None) -> str:
    label = STATUS_LABELS.get(status, status)
    if pending_close:
        label += " (待確認關單)"
    return label


def _state_cell_text(cell: dict | None,
                     issue_week_year: int,
                     issue_week_number: int) -> str:
    """Three-line cell content: Label / YYYY-MM-DD / short_note. Empty lines
    elided so a state-only cell doesn't have trailing blanks."""
    if not cell or not cell["state"]:
        return ""
    sty = STATE_STYLES.get(cell["state"])
    label = sty["label"] if sty else cell["state"]
    parts = [label]
    formatted = format_check_in_date(
        cell["check_in_date"], issue_week_year, issue_week_number
    )
    if formatted:
        parts.append(formatted)
    short_note = cell["short_note"] if "short_note" in cell.keys() else None
    if short_note:
        parts.append(short_note)
    return "\n".join(parts)


# ─── Workbook builder ──────────────────────────────────────────────────

def build_workbook(*, ongoing_issues, on_hold_issues, closed_issues,
                   nodes, all_states,
                   red_line_year, red_line_week,
                   exporter_display_name, exporter_username,
                   filtered: bool,
                   gitea_url_for=None) -> BytesIO:
    """Build the .xlsx into a BytesIO. Caller is responsible for sending it.

    `gitea_url_for` is the helper that turns a display_number into a Gitea URL;
    pass None to skip hyperlinks (useful for tests).
    """
    wb = Workbook()
    apply_taiwan_theme(wb)

    # Sheet 1: Ongoing (incl. on_hold mixed in).
    ws_ongoing = wb.active
    ws_ongoing.title = "Ongoing"
    _populate_sheet(
        ws_ongoing,
        title="Ongoing + On Hold",
        issues=list(ongoing_issues) + list(on_hold_issues),
        nodes=nodes,
        all_states=all_states,
        red_line_year=red_line_year,
        red_line_week=red_line_week,
        exporter_display_name=exporter_display_name,
        filtered=filtered,
        gitea_url_for=gitea_url_for,
    )

    # Sheet 2: Closed.
    ws_closed = wb.create_sheet("Closed")
    _populate_sheet(
        ws_closed,
        title="Closed",
        issues=list(closed_issues),
        nodes=nodes,
        all_states=all_states,
        red_line_year=red_line_year,
        red_line_week=red_line_week,
        exporter_display_name=exporter_display_name,
        filtered=filtered,
        gitea_url_for=gitea_url_for,
    )

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _populate_sheet(ws, *, title, issues, nodes, all_states,
                    red_line_year, red_line_week,
                    exporter_display_name, filtered, gitea_url_for):
    # ── Metadata band (rows 1-2) ─────────────────────────────────────
    rl_label = (f"wk{red_line_year - 2020}{red_line_week:02d}"
                if red_line_year and red_line_week else "—")
    now_str = datetime.now(timezone.utc).astimezone().strftime("%Y-%m-%d %H:%M")
    scope = "目前篩選結果" if filtered else "全部"
    meta_text = (
        f"{title} | 匯出時間 {now_str} | 匯出者 {exporter_display_name} | "
        f"紅線 {rl_label} | 範圍 {scope} | 題目數 {len(issues)}"
    )
    ws.cell(row=1, column=1, value=meta_text).font = metadata_font()

    legend_text = "圖例：  " + "  ·  ".join(
        f"{STATE_STYLES[s]['label']}" for s in
        ("done", "uat_done", "uat", "tbd", "developing", "unneeded")
    ) + "    紅線以上 UAT/TBD 以紅字標記"
    legend_cell = ws.cell(row=2, column=1, value=legend_text)
    legend_cell.font = metadata_font()

    # ── Header (row 3) ───────────────────────────────────────────────
    headers = [h for h, _ in FIXED_LEAD]
    headers += [n["display_name"] for n in nodes]
    headers += [h for h, _ in FIXED_TAIL]

    HEADER_ROW = 3
    for col_idx, label in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=col_idx, value=label)
        cell.font = header_font()
        cell.border = header_border()
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    widths = [w for _, w in FIXED_LEAD] + [NODE_WIDTH] * len(nodes) + [w for _, w in FIXED_TAIL]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Body rows, grouped by week with a wkXXX separator ────────────
    total_cols = len(headers)
    row_idx = HEADER_ROW + 1
    current_week = None
    prev_above_red_line = None
    for issue in issues:
        wk_key = (issue["week_year"], issue["week_number"])
        issue_above = above_red_line(wk_key[0], wk_key[1],
                                     red_line_year, red_line_week)

        if wk_key != current_week:
            # Red-line marker: insert between the last "above" wk-group and the
            # first "below" wk-group, so the divider sits right where the
            # threshold falls.
            if (red_line_year and red_line_week
                    and prev_above_red_line is True
                    and issue_above is False):
                _write_red_line_row(ws, row_idx, total_cols,
                                    red_line_year, red_line_week)
                row_idx += 1

            _write_wk_separator(ws, row_idx, wk_key, total_cols)
            current_week = wk_key
            row_idx += 1

        _write_issue_row(ws, row_idx, issue, nodes, all_states,
                         red_line_year, red_line_week, gitea_url_for)
        row_idx += 1
        prev_above_red_line = issue_above

    # ── Frozen panes: header + first 3 cols ──────────────────────────
    # Freeze at D{HEADER_ROW+1} → rows 1..HEADER_ROW frozen at top, cols A..C
    # frozen on the left.
    ws.freeze_panes = ws.cell(row=HEADER_ROW + 1, column=4)


def _write_wk_separator(ws, row_idx, wk_key, total_cols):
    """Light-gray band across the whole row with the wk label in col 1."""
    wk_label = f"wk{wk_key[0] - 2020}{wk_key[1]:02d}"
    fill = week_separator_fill()
    border = week_separator_border()
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row_idx, column=col,
                       value=wk_label if col == 1 else None)
        cell.fill = fill
        cell.border = border
        if col == 1:
            cell.font = week_separator_font()


def _write_red_line_row(ws, row_idx, total_cols, red_line_year, red_line_week):
    """Red-tinted band marking the red-line boundary. Spans all columns; label
    sits in col 1 so it's visible in narrow viewports too."""
    rl_label = f"wk{red_line_year - 2020}{red_line_week:02d}"
    label = f"─── redline {rl_label} ───"
    fill = red_line_row_fill()
    font = red_line_row_font()
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=row_idx, column=col,
                       value=label if col == 1 else None)
        cell.fill = fill
        if col == 1:
            cell.font = font


def _write_issue_row(ws, row_idx, issue, nodes, all_states,
                     red_line_year, red_line_week, gitea_url_for):
    above = above_red_line(issue["week_year"], issue["week_number"],
                           red_line_year, red_line_week)

    # Col 1: # (display_number) with hyperlink
    dn_cell = ws.cell(row=row_idx, column=1, value=issue["display_number"])
    if gitea_url_for:
        url = (gitea_url_for(issue["display_number"]) or "").strip()
        if url:
            dn_cell.hyperlink = Hyperlink(ref=dn_cell.coordinate, target=url)
            dn_cell.font = hyperlink_font()
        else:
            dn_cell.font = body_font_default()
    else:
        dn_cell.font = body_font_default()
    dn_cell.alignment = Alignment(horizontal="right", vertical="top")

    # Col 2: Status (friendly label, with pending_close suffix)
    status_cell = ws.cell(row=row_idx, column=2,
                          value=_status_label(issue["status"], issue["pending_close"]))
    status_cell.font = body_font_default()
    status_cell.alignment = Alignment(horizontal="left", vertical="top")

    # Col 3: Owner (= requestor_name; matches tracker UI)
    owner_cell = ws.cell(row=row_idx, column=3, value=issue["requestor_name"] or "")
    owner_cell.font = body_font_default()
    owner_cell.alignment = Alignment(horizontal="left", vertical="top")

    # Cols 4..(3+len(nodes)): per-node state cells
    states = all_states.get(issue["id"], {})
    for i, node in enumerate(nodes):
        col_idx = 4 + i
        cell_data = states.get(node["id"])
        text = _state_cell_text(cell_data, issue["week_year"], issue["week_number"])
        ws_cell = ws.cell(row=row_idx, column=col_idx, value=text)
        font = state_font(
            cell_data["state"] if cell_data else None,
            above_red_line=above,
        )
        if font:
            ws_cell.font = font
        else:
            ws_cell.font = body_font_default()
        ws_cell.alignment = Alignment(
            horizontal="center", vertical="top", wrap_text=True
        )

    # Tail columns: JIRA, UAT Path, Topic, ICV, Group Label
    base = 3 + len(nodes)
    tail_values = [
        issue["jira_ticket"] or "",
        issue["uat_path"] or "",
        issue["topic"] or "",
        issue["icv"] or "",
        issue["group_label"] or "",
    ]
    for offset, value in enumerate(tail_values, start=1):
        c = ws.cell(row=row_idx, column=base + offset, value=value)
        c.font = body_font_default()
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
