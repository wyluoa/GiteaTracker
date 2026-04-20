"""
Main routes — dashboard, tracker view, calendar, closed, search/filter, mark read, export.
"""
import calendar as cal_mod
from datetime import date, datetime, timezone, timedelta
from io import BytesIO
from pathlib import Path

from flask import (
    Blueprint, render_template, request, redirect, url_for, flash, g,
    send_file, send_from_directory, abort,
)

from app.db import get_db
from app.routes.auth import login_required, super_user_required, manager_or_super_required, optional_login
from app.models import issue as issue_model
from app.models import node as node_model
from app.models import issue_node_state as state_model
from app.models import setting as setting_model
from app.models import user as user_model
from app.models import joke as joke_model
from app.models import feedback as feedback_model

bp = Blueprint("main", __name__)


# ── Landing page → redirect to dashboard ──

@bp.route("/")
def index():
    """首頁 — 重導至 Tracker
    ---
    tags:
      - Tracker
    responses:
      302:
        description: 重導至 /tracker
    """
    return redirect(url_for("main.tracker"))


# ── Dashboard ──

@bp.route("/dashboard")
@manager_or_super_required
def dashboard():
    """儀表板 — 各 Node 卡片、趨勢圖表、Insights
    ---
    tags:
      - Dashboard
    responses:
      200:
        description: 渲染 dashboard.html，包含統計卡片、趨勢圖表、瓶頸分析
    """
    nodes = node_model.get_all_active()
    red_line_year, red_line_week = setting_model.get_red_line()

    node_counts = issue_model.dashboard_node_counts(red_line_year, red_line_week)
    ready_issues = issue_model.list_ready_to_close()
    ready_count = len(ready_issues)
    pending_close_issues = issue_model.list_pending_close()
    pending_close_count = len(pending_close_issues)
    on_hold_count = issue_model.count_by_status("on_hold")
    ongoing_count = issue_model.count_by_status("ongoing")
    closed_count = issue_model.count_closed()

    # Trend data for charts (auto-calculated)
    trends = issue_model.get_dashboard_trends()

    # Manual trend data (from Admin > Trend Data)
    db = get_db()
    manual_trend_rows = db.execute(
        "SELECT * FROM weekly_trend_data ORDER BY week_year, week_number"
    ).fetchall()
    manual_trend = {
        "weeks": [],
        "data": [],
    }
    for r in manual_trend_rows:
        manual_trend["weeks"].append(f"wk{r['week_year'] - 2020}{r['week_number']:02d}")
        total = r["cnt_uat"] + r["cnt_tbd"] + r["cnt_dev"] + r["cnt_close"]
        rate = round(r["cnt_close"] / total * 100, 1) if total else 0
        manual_trend["data"].append({
            "UAT": r["cnt_uat"], "TBD": r["cnt_tbd"],
            "Dev": r["cnt_dev"], "Close": r["cnt_close"],
            "rate": rate,
        })
    total_all = ongoing_count + on_hold_count + closed_count
    closing_rate = round(closed_count / total_all * 100, 1) if total_all else 0

    # Per-node UAT / TBD / Dev counts
    uat_total, uat_per_node = issue_model.count_node_states_by_type("uat")
    tbd_total, tbd_per_node = issue_model.count_node_states_by_type("tbd")
    dev_total, dev_per_node = issue_model.count_node_states_by_type("developing")

    # Closing rate excluding MtM
    closing_rate_ex_mtm, eff_closed_ex_mtm, _ = issue_model.closing_rate_excluding_node("n_mtm")

    # Insight data
    bottleneck = issue_model.get_bottleneck_nodes()
    velocity = issue_model.get_weekly_velocity()
    aging = issue_model.get_aging_stats()
    almost_done = issue_model.get_almost_done_issues(max_remaining=2)

    return render_template(
        "dashboard.html",
        nodes=nodes,
        node_counts=node_counts,
        ready_count=ready_count,
        ready_issues=ready_issues,
        pending_close_count=pending_close_count,
        pending_close_issues=pending_close_issues,
        on_hold_count=on_hold_count,
        ongoing_count=ongoing_count,
        closed_count=closed_count,
        closing_rate=closing_rate,
        trends=trends,
        manual_trend=manual_trend,
        uat_total=uat_total, uat_per_node=uat_per_node,
        tbd_total=tbd_total, tbd_per_node=tbd_per_node,
        dev_total=dev_total, dev_per_node=dev_per_node,
        closing_rate_ex_mtm=closing_rate_ex_mtm,
        eff_closed_ex_mtm=eff_closed_ex_mtm,
        bottleneck=bottleneck,
        velocity=velocity,
        aging=aging,
        almost_done=almost_done,
        red_line_year=red_line_year,
        red_line_week=red_line_week,
        user=g.current_user,
    )


# ── Tracker (main table) ──

@bp.route("/tracker")
@optional_login
def tracker():
    """追蹤器主表 — 搜尋 / 篩選 / 進階篩選
    ---
    tags:
      - Tracker
    parameters:
      - name: q
        in: query
        type: string
        description: 文字搜尋 (比對 #號、Topic、JIRA、UAT Path)
      - name: owner
        in: query
        type: string
        description: 篩選 Owner (requestor_name)
      - name: state
        in: query
        type: string
        description: 篩選狀態 (done/uat_done/uat/developing/tbd/unneeded/__blank__)
      - name: week_from
        in: query
        type: string
        description: 起始週 (格式 YYYYWW，如 202601)
      - name: week_to
        in: query
        type: string
        description: 結束週 (格式 YYYYWW，如 202612)
      - name: adv_node_1
        in: query
        type: string
        description: 進階篩選 1 — Node ID
      - name: adv_state_1
        in: query
        type: string
        description: 進階篩選 1 — 狀態
    responses:
      200:
        description: 渲染 tracker.html，依週分組顯示 ongoing 和 on_hold issues
    """
    nodes = node_model.get_all_active()

    # ── Search & filter params ──
    q = request.args.get("q", "").strip()
    filter_owner = request.args.get("owner", "").strip()
    filter_node = request.args.get("node", "").strip()
    filter_state = request.args.get("state", "").strip()
    filter_week_from = request.args.get("week_from", "").strip()
    filter_week_to = request.args.get("week_to", "").strip()

    # ── Advanced filter: up to 3 node+state pairs ──
    adv_filters = []
    for i in range(1, 4):
        anode = request.args.get(f"adv_node_{i}", "").strip()
        astate = request.args.get(f"adv_state_{i}", "").strip()
        if anode or astate:
            adv_filters.append({"node": anode, "state": astate, "index": i})
    # Match mode: 'all' = AND between conditions, 'any' = OR
    adv_match = request.args.get("adv_match", "all").strip().lower()
    if adv_match not in ("all", "any"):
        adv_match = "all"

    ongoing_issues = issue_model.get_ongoing()
    on_hold_issues = issue_model.get_on_hold()

    has_basic_filter = (q or filter_owner or filter_node or filter_state
                        or filter_week_from or filter_week_to)

    # Apply basic filters (text, owner, week range)
    if has_basic_filter:
        ongoing_issues = _apply_filters(ongoing_issues, nodes, q, filter_owner, filter_state,
                                         filter_week_from, filter_week_to)
        on_hold_issues = _apply_filters(on_hold_issues, nodes, q, filter_owner, filter_state,
                                         filter_week_from, filter_week_to)

    # Bulk load all node states
    all_issue_ids = [i["id"] for i in ongoing_issues] + [i["id"] for i in on_hold_issues]
    all_states = state_model.get_all_states_for_issues(all_issue_ids)

    # Basic Node + State filter (one-row equivalent of advanced filter).
    #   node only  → that node has any state set
    #   state only → any node matches the state
    #   both       → that specific node's state equals filter
    basic_node_id = int(filter_node) if filter_node.isdigit() else None
    if (basic_node_id or filter_state) and all_states:
        filtered_ids = set()
        for issue_id, node_states in all_states.items():
            if basic_node_id and filter_state:
                cell = node_states.get(basic_node_id)
                if filter_state == "__blank__":
                    if not cell or not cell["state"]:
                        filtered_ids.add(issue_id)
                elif cell and cell["state"] == filter_state:
                    filtered_ids.add(issue_id)
            elif basic_node_id:
                cell = node_states.get(basic_node_id)
                if cell and cell["state"]:
                    filtered_ids.add(issue_id)
            elif filter_state:
                for nid, cell in node_states.items():
                    if filter_state == "__blank__":
                        if not cell["state"]:
                            filtered_ids.add(issue_id)
                            break
                    elif cell["state"] == filter_state:
                        filtered_ids.add(issue_id)
                        break
        ongoing_issues = [i for i in ongoing_issues if i["id"] in filtered_ids]
        on_hold_issues = [i for i in on_hold_issues if i["id"] in filtered_ids]
        all_issue_ids = [i["id"] for i in ongoing_issues] + [i["id"] for i in on_hold_issues]
        all_states = state_model.get_all_states_for_issues(all_issue_ids)

    # Apply advanced node+state filters.
    #   adv_match = 'all' → intersect (AND), 'any' → union (OR)
    if adv_filters:
        all_current_ids = set(i["id"] for i in ongoing_issues) | set(i["id"] for i in on_hold_issues)

        def _matches(issue_id, af_node_id, af_state):
            node_states = all_states.get(issue_id, {})
            if af_node_id and af_state:
                cell = node_states.get(af_node_id)
                if af_state == "__blank__":
                    return not cell or not cell["state"]
                return bool(cell and cell["state"] == af_state)
            if af_node_id and not af_state:
                cell = node_states.get(af_node_id)
                return bool(cell and cell["state"])
            if not af_node_id and af_state:
                for _, cell in node_states.items():
                    if af_state == "__blank__":
                        if not cell["state"]:
                            return True
                    elif cell["state"] == af_state:
                        return True
                return False
            return False

        per_condition_ids = []
        for af in adv_filters:
            af_node_id = int(af["node"]) if af["node"] else None
            af_state = af["state"]
            matched = {iid for iid in all_current_ids if _matches(iid, af_node_id, af_state)}
            per_condition_ids.append(matched)

        if adv_match == "any":
            kept = set().union(*per_condition_ids)
        else:  # 'all'
            kept = set.intersection(*per_condition_ids)

        ongoing_issues = [i for i in ongoing_issues if i["id"] in kept]
        on_hold_issues = [i for i in on_hold_issues if i["id"] in kept]
        all_issue_ids = [i["id"] for i in ongoing_issues] + [i["id"] for i in on_hold_issues]
        all_states = state_model.get_all_states_for_issues(all_issue_ids)

    # Separate issues with group_label from week-based issues
    week_issues = [i for i in ongoing_issues if not i["group_label"]]
    labeled_issues = [i for i in ongoing_issues if i["group_label"]]

    # Group week-based issues by week
    week_groups = []
    current_key = None
    current_group = None
    for issue in week_issues:
        key = (issue["week_year"], issue["week_number"])
        if key != current_key:
            current_key = key
            current_group = {"week_year": key[0], "week_number": key[1], "issues": []}
            week_groups.append(current_group)
        current_group["issues"].append(issue)

    # Group labeled issues by group_label
    label_groups = []
    current_label = None
    current_lgroup = None
    for issue in labeled_issues:
        if issue["group_label"] != current_label:
            current_label = issue["group_label"]
            current_lgroup = {"label": current_label, "issues": []}
            label_groups.append(current_lgroup)
        current_lgroup["issues"].append(issue)

    # Red line
    red_line_year, red_line_week = setting_model.get_red_line()

    # Current ISO week
    today = date.today()
    iso = today.isocalendar()

    # Version diff: last_viewed_at
    last_viewed = g.current_user["last_viewed_at"] if g.current_user else None

    # Distinct owners for filter dropdown
    db = get_db()
    owners = db.execute(
        "SELECT DISTINCT requestor_name FROM issues WHERE requestor_name IS NOT NULL AND is_deleted = 0 ORDER BY requestor_name"
    ).fetchall()

    return render_template(
        "tracker.html",
        nodes=nodes,
        week_groups=week_groups,
        label_groups=label_groups,
        on_hold_issues=on_hold_issues,
        all_states=all_states,
        red_line_year=red_line_year,
        red_line_week=red_line_week,
        current_week_year=iso[0],
        current_week_number=iso[1],
        last_viewed=last_viewed,
        user=dict(g.current_user) if g.current_user else {},
        # Filter state
        q=q, filter_owner=filter_owner, filter_node=filter_node,
        filter_state=filter_state,
        filter_week_from=filter_week_from, filter_week_to=filter_week_to,
        owners=[r["requestor_name"] for r in owners],
        adv_filters=adv_filters,
        adv_match=adv_match,
    )


def _apply_filters(issues, nodes, q, owner, state, week_from, week_to):
    result = issues
    if q:
        ql = q.lower()
        result = [i for i in result if
                  ql in (i["display_number"] or "").lower() or
                  ql in (i["topic"] or "").lower() or
                  ql in (i["jira_ticket"] or "").lower() or
                  ql in (i["uat_path"] or "").lower()]
    if owner:
        result = [i for i in result if i["requestor_name"] == owner]
    if week_from:
        try:
            wy, wn = int(week_from[:4]), int(week_from[4:])
            result = [i for i in result if (i["week_year"], i["week_number"]) >= (wy, wn)]
        except (ValueError, IndexError):
            pass
    if week_to:
        try:
            wy, wn = int(week_to[:4]), int(week_to[4:])
            result = [i for i in result if (i["week_year"], i["week_number"]) <= (wy, wn)]
        except (ValueError, IndexError):
            pass
    return result


# ── Mark all as read ──

@bp.route("/mark_all_read", methods=["POST"])
@login_required
def mark_all_read():
    """標記全部已讀
    ---
    tags:
      - Tracker
    responses:
      302:
        description: 更新 last_viewed_at 後重導至 tracker
    """
    user_model.update_last_viewed(g.current_user["id"])
    flash("已標記全部為已讀", "success")
    return redirect(url_for("main.tracker"))


# ── Export Excel ──

@bp.route("/export")
@login_required
def export_excel():
    """匯出 Excel — 下載目前所有 ongoing + on_hold issues
    ---
    tags:
      - Export
    produces:
      - application/vnd.openxmlformats-officedocument.spreadsheetml.sheet
    responses:
      200:
        description: 下載 .xlsx 檔案，包含 #號、狀態、Owner、各 Node 狀態、JIRA、UAT Path、Topic
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    nodes = node_model.get_all_active()
    ongoing = issue_model.get_ongoing()
    on_hold = issue_model.get_on_hold()
    all_issues = list(ongoing) + list(on_hold)
    all_ids = [i["id"] for i in all_issues]
    all_states = state_model.get_all_states_for_issues(all_ids)

    STATE_LABELS = {
        "done": "Done", "uat_done": "UAT done", "uat": "UAT",
        "developing": "Developing", "tbd": "TBD", "unneeded": "Unneeded",
    }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ongoing"

    # Header
    headers = ["#", "Status", "Owner"] + [n["display_name"] for n in nodes] + ["JIRA", "UAT Path", "Topic"]
    ws.append(headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    current_week = None
    for issue in all_issues:
        wk = (issue["week_year"], issue["week_number"])
        if wk != current_week:
            current_week = wk
            ws.append([f"wk{wk[0] - 2020}{wk[1]:02d}"])

        row_data = [issue["display_number"], issue["status"], issue["requestor_name"] or ""]
        states = all_states.get(issue["id"], {})
        for node in nodes:
            cell = states.get(node["id"])
            if cell and cell["state"]:
                label = STATE_LABELS.get(cell["state"], cell["state"])
                if cell["check_in_date"]:
                    label += f"\n{cell['check_in_date']}"
                row_data.append(label)
            else:
                row_data.append("")
        row_data.extend([issue["jira_ticket"] or "", issue["uat_path"] or "", issue["topic"]])
        ws.append(row_data)

    # Auto-width
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"gitea_tracker_{date.today().isoformat()}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)


# ── Calendar ──

def _parse_check_in_date(raw, today):
    """Parse a stored check_in_date to a date object.

    DB stores two formats: ISO "YYYY-MM-DD" (from UI date pickers) and
    short "MM-DD" (from Excel import, no year). For the short form we
    assume today's year, unless the resulting date is more than 180 days
    in the future — in that case it probably came from last year.
    Returns None if parsing fails.
    """
    if not raw:
        return None
    s = str(raw).strip()
    try:
        if len(s) == 10 and s[4] == '-' and s[7] == '-':
            return datetime.strptime(s, "%Y-%m-%d").date()
        if len(s) == 5 and s[2] == '-':
            d = datetime.strptime(f"{today.year}-{s}", "%Y-%m-%d").date()
            if (d - today).days > 180:
                d = d.replace(year=today.year - 1)
            return d
    except ValueError:
        return None
    return None


@bp.route("/calendar")
@manager_or_super_required
def calendar_view():
    """行事曆 — 依月份顯示 check-in 日期
    ---
    tags:
      - Calendar
    parameters:
      - name: year
        in: query
        type: integer
        description: 年份 (預設當年)
      - name: month
        in: query
        type: integer
        description: 月份 (預設當月)
    responses:
      200:
        description: 渲染月曆，顯示各 cell 的 check-in 日期
    """
    today = date.today()
    year = request.args.get("year", today.year, type=int)
    month = request.args.get("month", today.month, type=int)

    # Clamp
    if month < 1:
        month, year = 12, year - 1
    elif month > 12:
        month, year = 1, year + 1

    # Date range for the month
    first_day = date(year, month, 1)
    last_day = date(year, month, cal_mod.monthrange(year, month)[1])

    db = get_db()
    rows = db.execute(
        """SELECT s.issue_id, s.node_id, s.state, s.check_in_date, s.short_note,
                  i.display_number, i.topic, n.display_name as node_name
           FROM issue_node_states s
           JOIN issues i ON s.issue_id = i.id
           JOIN nodes n ON s.node_id = n.id
           WHERE s.check_in_date BETWEEN ? AND ?
             AND i.status = 'ongoing' AND i.is_deleted = 0
           ORDER BY s.check_in_date, i.display_number""",
        (first_day.isoformat(), last_day.isoformat()),
    ).fetchall()

    # Group by date string
    events_by_date = {}
    for r in rows:
        events_by_date.setdefault(r["check_in_date"], []).append(r)

    # Build calendar grid
    cal = cal_mod.Calendar(firstweekday=0)  # Monday first
    weeks = cal.monthdatescalendar(year, month)

    # Prev/next month
    prev_month = month - 1 if month > 1 else 12
    prev_year = year if month > 1 else year - 1
    next_month = month + 1 if month < 12 else 1
    next_year = year if month < 12 else year + 1

    nodes = node_model.get_all_active()

    # ── Overdue-to-launch list (super user only) ──
    # Nodes whose check_in_date has arrived or passed but the state isn't
    # done/unneeded. Helps developers spot items that *should* have gone live
    # but haven't. Handles both DB formats: YYYY-MM-DD (UI) and MM-DD (Excel).
    overdue = []
    if g.current_user and g.current_user["is_super_user"]:
        overdue_rows = db.execute(
            """SELECT s.issue_id, s.node_id, s.state, s.check_in_date, s.short_note,
                      i.display_number, i.topic, i.requestor_name,
                      n.display_name as node_name
               FROM issue_node_states s
               JOIN issues i ON s.issue_id = i.id
               JOIN nodes n ON s.node_id = n.id
               WHERE s.check_in_date IS NOT NULL AND s.check_in_date != ''
                 AND (s.state IS NULL OR s.state NOT IN ('done', 'unneeded'))
                 AND i.status = 'ongoing' AND i.is_deleted = 0"""
        ).fetchall()

        for r in overdue_rows:
            parsed = _parse_check_in_date(r["check_in_date"], today)
            if parsed and parsed <= today:
                overdue.append({
                    "issue_id": r["issue_id"],
                    "node_id": r["node_id"],
                    "display_number": r["display_number"],
                    "topic": r["topic"],
                    "requestor_name": r["requestor_name"],
                    "node_name": r["node_name"],
                    "state": r["state"],
                    "short_note": r["short_note"],
                    "check_in_date": parsed.isoformat(),
                    "days_overdue": (today - parsed).days,
                })
        overdue.sort(key=lambda x: (-x["days_overdue"], x["display_number"]))

    return render_template(
        "calendar.html",
        year=year, month=month,
        weeks=weeks,
        events_by_date=events_by_date,
        today=today,
        prev_year=prev_year, prev_month=prev_month,
        next_year=next_year, next_month=next_month,
        nodes=nodes,
        overdue=overdue,
        user=g.current_user,
    )


# ── Closed Issues ──

@bp.route("/closed")
@optional_login
def closed():
    """已關單列表 — 分頁 + 搜尋
    ---
    tags:
      - Tracker
    parameters:
      - name: page
        in: query
        type: integer
        description: 頁碼 (預設 1)
      - name: q
        in: query
        type: string
        description: 搜尋 (比對 #號、Topic、JIRA)
    responses:
      200:
        description: 渲染 closed.html，每頁 50 筆
    """
    page = request.args.get("page", 1, type=int)
    q = request.args.get("q", "").strip()
    per_page = 50

    db = get_db()

    if q:
        ql = f"%{q}%"
        rows = db.execute(
            """SELECT * FROM issues
               WHERE status = 'closed' AND is_deleted = 0
                 AND (display_number LIKE ? OR topic LIKE ? OR jira_ticket LIKE ?)
               ORDER BY week_year, week_number, CAST(display_number AS INTEGER)
               LIMIT ? OFFSET ?""",
            (ql, ql, ql, per_page, (page - 1) * per_page),
        ).fetchall()
        total_row = db.execute(
            """SELECT COUNT(*) as cnt FROM issues
               WHERE status = 'closed' AND is_deleted = 0
                 AND (display_number LIKE ? OR topic LIKE ? OR jira_ticket LIKE ?)""",
            (ql, ql, ql),
        ).fetchone()
        total = total_row["cnt"]
    else:
        total = issue_model.count_closed()
        rows = db.execute(
            """SELECT * FROM issues
               WHERE status = 'closed' AND is_deleted = 0
               ORDER BY week_year, week_number, CAST(display_number AS INTEGER)
               LIMIT ? OFFSET ?""",
            (per_page, (page - 1) * per_page),
        ).fetchall()

    # Split week-based issues from labeled (group_label) issues
    week_rows = [i for i in rows if not i["group_label"]]
    labeled_rows = sorted(
        (i for i in rows if i["group_label"]),
        key=lambda i: (i["group_label"], i["week_year"] or 0, i["week_number"] or 0),
    )

    # Group week-based by (week_year, week_number)
    week_groups = []
    current_key = None
    current_group = None
    for issue in week_rows:
        key = (issue["week_year"], issue["week_number"])
        if key != current_key:
            current_key = key
            current_group = {"week_year": key[0], "week_number": key[1], "issues": []}
            week_groups.append(current_group)
        current_group["issues"].append(issue)

    # Group labeled issues by group_label
    label_groups = []
    current_label = None
    current_lgroup = None
    for issue in labeled_rows:
        if issue["group_label"] != current_label:
            current_label = issue["group_label"]
            current_lgroup = {"label": current_label, "issues": []}
            label_groups.append(current_lgroup)
        current_lgroup["issues"].append(issue)

    nodes = node_model.get_all_active()
    all_ids = [i["id"] for i in rows]
    all_states = state_model.get_all_states_for_issues(all_ids)
    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template(
        "closed.html",
        week_groups=week_groups,
        label_groups=label_groups,
        nodes=nodes,
        all_states=all_states,
        page=page,
        total_pages=total_pages,
        total=total,
        q=q,
        user=dict(g.current_user) if g.current_user else {},
    )


@bp.route("/healthz")
def healthz():
    """健康檢查
    ---
    tags:
      - Health
    responses:
      200:
        description: '回傳 {"status": "ok"}'
        schema:
          type: object
          properties:
            status:
              type: string
              example: ok
    """
    return {"status": "ok"}


# ── In-app documentation (static HTML under docs/) ──

_DOCS_DIR = Path(__file__).resolve().parent.parent.parent / "docs"

# filename + role required. "user" = any logged-in user, "super" = super_user only.
_GUIDES = {
    "user":   ("user_guide.html",        "user"),
    "dev":    ("developer_guide.html",   "super"),
    "report": ("management_report.html", "super"),
}


@bp.route("/guide/<name>")
@login_required
def guide(name):
    """使用者手冊 / 開發者指南 / 管理報告 — 提供內部 HTML 文件頁面
    ---
    tags:
      - Docs
    parameters:
      - name: name
        in: path
        type: string
        required: true
        description: user | dev | report (dev/report require super user)
    responses:
      200:
        description: 回傳對應的 HTML 文件頁面
      403:
        description: 權限不足（dev / report 需 super user）
      404:
        description: 名稱不存在
    """
    entry = _GUIDES.get(name)
    if not entry:
        abort(404)
    filename, role = entry
    if role == "super" and not g.current_user["is_super_user"]:
        abort(403)
    return send_from_directory(_DOCS_DIR, filename)


# ── Easter egg: /fun (jokes / light stories for meeting warm-ups) ──

@bp.route("/fun", methods=["GET", "POST"])
@login_required
def fun():
    """笑話 / 小品頁 — 會議開場暖場用 (彩蛋入口: 連點 navbar brand 5 次)
    ---
    tags:
      - Fun
    responses:
      200:
        description: 渲染 fun.html，顯示所有笑話
    """
    if request.method == "POST":
        # Only super_user may add (mode b: you curate)
        if not g.current_user["is_super_user"]:
            abort(403)
        body = request.form.get("body", "").strip()
        if body:
            joke_model.create(
                body=body,
                author_user_id=g.current_user["id"],
                author_name_snapshot=g.current_user["display_name"],
            )
            flash("新增完成 🎉", "success")
        return redirect(url_for("main.fun"))

    jokes = joke_model.list_all()
    return render_template("fun.html", jokes=jokes, user=g.current_user)


@bp.route("/fun/<int:joke_id>/delete", methods=["POST"])
@super_user_required
def fun_delete(joke_id):
    """刪除一則笑話 (super user only, soft delete)"""
    joke_model.soft_delete(joke_id)
    flash("已刪除", "success")
    return redirect(url_for("main.fun"))


@bp.route("/fun/random")
@login_required
def fun_random():
    """回傳一則隨機笑話的 HTML partial (for HTMX swap)"""
    joke = joke_model.get_random()
    return render_template("partials/joke_card.html", joke=joke, user=g.current_user)


# ── Feedback (all logged-in users can submit; super user reviews in Admin) ──

@bp.route("/feedback", methods=["GET", "POST"])
@login_required
def feedback():
    """意見回饋 — 使用者提交 bug / feature / other；看自己以前送出的 + admin 回覆
    ---
    tags:
      - Feedback
    parameters:
      - name: category
        in: formData
        type: string
        description: bug / feature / other
      - name: body
        in: formData
        type: string
        description: 回饋內容
    responses:
      200:
        description: 顯示送出表單 + 自己以前的紀錄
      302:
        description: 送出後重導回 /feedback
    """
    if request.method == "POST":
        category = (request.form.get("category") or "").strip()
        body = (request.form.get("body") or "").strip()
        if not body:
            flash("請填入內容", "error")
            return redirect(url_for("main.feedback"))
        feedback_model.create(
            author_user_id=g.current_user["id"],
            author_name_snapshot=g.current_user["display_name"],
            category=category,
            body=body,
        )
        flash("已送出，謝謝你的回饋！", "success")
        return redirect(url_for("main.feedback"))

    items = feedback_model.list_by_author(g.current_user["id"])
    return render_template(
        "feedback.html",
        items=items,
        categories=feedback_model.VALID_CATEGORIES,
        user=g.current_user,
    )
