"""
Shared Excel parsing utilities.

Used by both the CLI import (import_from_excel.py) and the web-based
Excel update feature (admin routes).
"""
import re

import openpyxl

# ---------- Constants ----------

STATE_MAP = {
    "done": "done",
    "uat done": "uat_done",
    "uatdone": "uat_done",
    "uat": "uat",
    "developing": "developing",
    "dev": "developing",
    "tbd": "tbd",
    "unneeded": "unneeded",
}

STATE_LABELS = {
    "done": "Done",
    "uat_done": "UAT Done",
    "uat": "UAT",
    "developing": "Developing",
    "tbd": "TBD",
    "unneeded": "Unneeded",
}

HEADER_TO_CODE = {
    "A10":   "n_a10",
    "A12":   "n_a12",
    "A14":   "n_a14",
    "N2":    "n_n2",
    "A16":   "n_a16",
    "N3":    "n_n3",
    "N4/N5": "n_n4n5",
    "N6/N7": "n_n6n7",
    "000":   "n_000",
    "MtM":   "n_mtm",
}

WK_PATTERN = re.compile(r"^wk(\d+)$", re.IGNORECASE)
DATE_PATTERN = re.compile(r"(\d{1,2})/(\d{1,2})")


# ---------- Cell parsing ----------

def parse_cell(raw_value):
    """Parse a cell value into (state, check_in_date, short_note).

    Examples:
        "Done"                         -> ("done", None, None)
        "UAT done\\n2/20 Check in"     -> ("uat_done", "02-20", None)
        "Developing added spec for N4" -> ("developing", None, "added spec for N4")
        ""  / None                     -> (None, None, None)
    """
    if raw_value is None:
        return None, None, None

    text = str(raw_value).strip()
    if not text:
        return None, None, None

    lines = [ln.strip() for ln in text.split("\n") if ln.strip()]
    if not lines:
        return None, None, None

    state = None
    check_in_date = None
    short_note = None

    first_line = lines[0]
    first_lower = first_line.lower()

    if first_lower in STATE_MAP:
        state = STATE_MAP[first_lower]
    else:
        for key in sorted(STATE_MAP.keys(), key=len, reverse=True):
            if first_lower.startswith(key):
                state = STATE_MAP[key]
                remainder = first_line[len(key):].strip()
                if remainder:
                    short_note = remainder
                break

    if state is None:
        short_note = text
        return None, None, short_note

    for line in lines[1:]:
        date_match = DATE_PATTERN.search(line)
        if date_match:
            month = int(date_match.group(1))
            day = int(date_match.group(2))
            check_in_date = f"{month:02d}-{day:02d}"
        else:
            if short_note:
                short_note += " " + line
            else:
                short_note = line

    return state, check_in_date, short_note


# ---------- Merged cells ----------

def expand_merged_cells(ws):
    """For each merged range, fill all cells with the top-left value."""
    for merge_range in list(ws.merged_cells.ranges):
        min_col, min_row, max_col, max_row = merge_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merge_range))
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value


# ---------- Sheet parsing (no DB writes) ----------

def parse_sheet(ws, node_lookup, is_closed_sheet=False):
    """Parse a worksheet into a list of issue dicts.

    Args:
        ws: openpyxl worksheet
        node_lookup: dict of code -> node_id
        is_closed_sheet: if True, all issues default to 'closed'

    Returns:
        list of dicts:
        {
            "display_number": "42",
            "topic": "...",
            "requestor_name": "...",
            "week_year": 2025,
            "week_number": 21,
            "jira_ticket": "...",
            "icv": "...",
            "uat_path": "...",
            "status": "ongoing",
            "nodes": {
                node_id: {"state": "uat", "check_in_date": "02-20", "short_note": "..."},
                ...
            }
        }
    """
    expand_merged_cells(ws)

    header_row = [cell.value for cell in ws[1]]

    node_columns = {}  # col_idx -> node_id
    owner_col = None
    status_col = None
    jira_col = None
    icv_col = None
    uat_path_col = None
    topic_col = None

    for idx, header in enumerate(header_row):
        if header is None:
            continue
        h = str(header).strip()
        if h == "Status":
            status_col = idx
        elif h == "Owner":
            owner_col = idx
        elif h == "JIRA":
            jira_col = idx
        elif h == "ICV":
            icv_col = idx
        elif h in ("UAT path", "UAT Path", "UATpath"):
            uat_path_col = idx
        elif h == "Topic":
            topic_col = idx
        elif h in HEADER_TO_CODE:
            code = HEADER_TO_CODE[h]
            if code in node_lookup:
                node_columns[idx] = node_lookup[code]

    if not node_columns:
        return []

    results = []
    current_week_year = None
    current_week_number = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        first_val = row[0].value
        if first_val is None:
            continue

        first_str = str(first_val).strip()
        if not first_str:
            continue

        wk_match = WK_PATTERN.match(first_str)
        if wk_match:
            raw_week = int(wk_match.group(1))
            if raw_week > 100:
                year_part = raw_week // 100
                week_part = raw_week % 100
                current_week_year = 2020 + year_part
                current_week_number = week_part
            else:
                current_week_year = current_week_year or 2025
                current_week_number = raw_week
            continue

        try:
            display_number = str(int(float(first_str)))
        except (ValueError, TypeError):
            continue

        if current_week_year is None:
            current_week_year = 2025
            current_week_number = 1

        def cell_val(idx):
            if idx is not None and idx < len(row):
                return row[idx].value
            return None

        status_val = str(cell_val(status_col) or "ongoing").strip().lower()
        issue_status = "ongoing"
        if is_closed_sheet or status_val == "closed":
            issue_status = "closed"
        elif status_val in ("on hold", "on_hold"):
            issue_status = "on_hold"

        owner_name = str(cell_val(owner_col) or "").strip() or None
        jira = str(cell_val(jira_col) or "").strip() or None
        icv = str(cell_val(icv_col) or "").strip() or None
        uat_path = str(cell_val(uat_path_col) or "").strip() or None
        topic = str(cell_val(topic_col) or "").strip() or f"Issue #{display_number}"

        nodes = {}
        for col_idx, node_id in node_columns.items():
            raw = cell_val(col_idx)
            state, check_in_date, short_note = parse_cell(raw)
            nodes[node_id] = {
                "state": state,
                "check_in_date": check_in_date,
                "short_note": short_note,
            }

        results.append({
            "display_number": display_number,
            "topic": topic,
            "requestor_name": owner_name,
            "week_year": current_week_year,
            "week_number": current_week_number,
            "jira_ticket": jira,
            "icv": icv,
            "uat_path": uat_path,
            "status": issue_status,
            "nodes": nodes,
        })

    return results


def parse_workbook(file_path, node_lookup):
    """Parse all sheets in a workbook.

    Returns a flat list of issue dicts from all sheets.
    """
    wb = openpyxl.load_workbook(file_path)
    all_issues = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_closed = "closed" in sheet_name.lower()
        issues = parse_sheet(ws, node_lookup, is_closed_sheet=is_closed)
        all_issues.extend(issues)
    return all_issues
